import os
from openpyxl import Workbook, load_workbook

file_name = "../sku_2.0.xlsx"
if not os.path.exists(file_name):
    raise FileNotFoundError(f"File '{file_name}' not found.")

p_door_shtuf = {
    "name": "p_door_shtuf",
    "verbose_name": "דלת שטוף",
    "sheet_name": "מוצרים לפי מידות",
    "sizes": [
        (65,203),
        (75,203),
        (85,203),
        (95,203),
        (110,203),
        (65,215),
        (75,215),
        (85,215),
        (95,215),
        (110,215),
        (120,242),
        (120,300)
    ]
}

# handles



def add_rows_to_excel(file, product_dict):
    workbook = load_workbook(file)
    ws = workbook[product_dict["sheet_name"]]
    start_row = ws.max_row + 1
    for i, (w, h) in enumerate(product_dict["sizes"], start=start_row):
        ws[f"A{i}"] = f"{product_dict['name']}_{w}x{h}"
        ws[f"B{i}"] = f"{product_dict['verbose_name']}_{w}x{h}"
    workbook.save(file_name)
    print(f"✅ Added new rows: {file_name}")


add_rows_to_excel(file_name, p_door_shtuf)