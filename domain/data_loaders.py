import openpyxl
from classes import *


class ExcelOrderLoader:
    """Implementation of OrderLoader for Excel files"""

    def __init__(self, file_path):
        self.file_path = file_path

    def load(self, order_num) -> Order:
        """Load order from the Excel file"""
        import openpyxl
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
        except FileNotFoundError as e:
            raise FileNotFoundError(f"Excel file not found: {self.file_path}") from e

        if str(order_num) not in wb.sheetnames:
            raise ValueError(f"Order sheet '{order_num}' not found in {self.file_path}")

        ws = wb[str(order_num)]
        order = Order(order_num)

        current_item = None
        customizers = get_all_customizers()
        item_num = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            pos_num = row[0]
            sku = row[1]
            if pos_num is not None:
                item_num = int(pos_num)
                current_item = None
                continue

            if sku is not None:
                if current_item is None:
                    width = row[3]
                    height = row[4]
                    direction = row[6]
                    opening = row[7]
                    undercut = row[8]
                    current_item = Item(item_num, sku, width, height, direction, opening, undercut)
                    order.items.append(current_item)
                else:
                    if sku in customizers:
                        current_item.customizers.append(customizers[sku])

        return order


def get_all_models_parameters():
    """Load model calculation parameters from sku_0.1.xlsx"""
    params = {}
    wb = openpyxl.load_workbook("sku_0.1.xlsx", data_only=True)
    ws = wb["Models"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = row[2]
        pcw = row[3]
        pch = row[4]
        ddc = row[5]
        if sku:
            params[sku] = {'pcw': pcw or 0, 'pch': pch or 0, 'ddc': ddc or 0}
    return params


def get_raw_bom_data():
    """Load all models and their components from BOM.xlsx into a dictionary"""
    raw_bom_data = {}
    wb = openpyxl.load_workbook("BOM.xlsx", data_only=True)
    ws = wb["models"]
    current_sku = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            current_sku = row[1]
            raw_bom_data[current_sku] = []
        else:
            raw_bom_data[current_sku].append(BOMItem(row[2], row[3], row[4]))
    return raw_bom_data

def get_raw_bom_components(sku: BOMItem):
    return get_raw_bom_data()[sku]

def get_bom_data():
    """create dictionary of full bom data"""
    raw_bom_data = get_raw_bom_data()
    full_bom_data = {}
    for sku, components in raw_bom_data.items():
        full_bom_data[sku] = []
        for component in components:
            if component.sku in raw_bom_data:
                full_bom_data[sku].extend(get_raw_bom_components(component.sku))
            else:
                full_bom_data[sku].append(component)
    return full_bom_data

def get_bom_components(sku: str):
    full_bom_data = get_bom_data()
    return full_bom_data[sku]

def get_sheet_materials():
    """Load sheet materials from sku_0.1.xlsx"""
    materials = {}
    wb = openpyxl.load_workbook("sku_0.1.xlsx", data_only=True)
    ws = wb["sheet_material"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[5]:  # SKU
            if row[6] not in materials:
                materials[row[6]] = []
            materials[row[6]].append(Material(row[5], row[1], row[2], row[3], row[4], row[6]))
    return materials


def get_all_customizers():
    wb = openpyxl.load_workbook("sku_0.1.xlsx", data_only=True)
    ws = wb["Customizers"]
    customizers = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        customizers[row[2]] = (Customizer(row[1], row[2], row[3], row[4], row[5], row[6]))
    return customizers

def get_all_component_changers():
    wb = openpyxl.load_workbook("sku_0.1.xlsx", data_only=True)
    ws = wb["Component Changers"]
    component_changers = []
    current_sku = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        compatible_models = row[3].split(",") if row[3] else []
        components_to_remove = [row[4]] if row[4] else []
        components_to_add = [(row[5], row[6], row[7])] if row[5] else []
        if row[1]:
            current_sku = row[1]
            component_changers.append(ComponentChanger(sku=row[1],
                                                               compatible_models=compatible_models,
                                                               components_to_remove=components_to_remove,
                                                               components_to_add=components_to_add))
        else:
            if components_to_add:
                component_changers[len(component_changers)-1].components_to_add.extend(components_to_add)
            if components_to_remove:
                component_changers[len(component_changers)-1].components_to_remove.extend(components_to_remove)


    return component_changers

# print(*get_all_component_changers(), sep="\n")
